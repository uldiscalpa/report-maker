from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side

def _as_text(value):
    if value is None:
        return ""
    return str(value)

class ExcelCreator:
    """excel report creator class"""

    format_as_float_list = [
        "Rentabilitāte",
        "Prāmja izmaksas",
        "Ienākumi",
        "Amorti. izm.",
        "Pārējie izdevumi",
        "Ienākumi",
        "Rēķina summa",
        "Peļņa",
    ]

    format_as_int_list = [
        "Degviela iztērēta",
        "Ienākumi",
        "Bruto alga",
    ]

    format_formula_as_sum = [
        "Ienākumi",
        "Degviela iztērēta",
        "Bruto alga",
        "Pārējie izdevumi",
        "Amorti. izm.",
        "Prāmja izmaksas",
        "Nobraukti km",
        "Peļņa"
    ]

    format_date_as_short_date = [
         "Reisa sākuma d.",
         "Reisa beigu d.",
    ]

    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def __init__(self, file_name='report.xlsx'):
        self.wb = Workbook()
        self.file_name = file_name
    
    def write_report(self, work_sheet_name= '', data_list= []):
        """Ekport data_list to init excel object"""
        ws = self.wb.create_sheet(work_sheet_name, 0)

        header_list= data_list[0].keys()
        self._header_writer(work_sheet= ws, header_list= list(header_list))
        self._data_writer(work_sheet= ws, data_list= data_list)
        self._table_column_summary_writer(work_sheet= ws)
        self._column_adjuster(work_sheet= ws)
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight = False
        ws.set_printer_settings(10, "landscape")
        self._save()
        
    def _column_adjuster(self, work_sheet):
        for column_cells in work_sheet.columns:
            length = max(len(_as_text(cell.value)) for cell in column_cells)
            print(column_cells[1].column_letter , length)
            # if length < 6: length= 7
            # work_sheet.column_dimensions[column_cells[0].column_letter].width= length
            # work_sheet.column_dimensions[column_cells[1].column_letter].bestFit = True
            work_sheet.column_dimensions[column_cells[1].column_letter].width = length / 1.2
            
    def _save(self):
        self.wb.save(filename=self.file_name)
        
    def _header_writer(self, work_sheet= None, header_list= []):
        for col, header in enumerate(header_list, start= 1):
            header_cell = work_sheet.cell(column=col, row= 1)
            header_cell.value = str(header).encode(encoding='UTF-8')
            header_cell.border = self.border
    
    def _data_writer(self, work_sheet= None, data_list= []):
        for row, data in enumerate(data_list, start= 2):
            col = 0
            for k, v in data.items():
                col += 1
                cell = work_sheet.cell(column=col, row=row)
                if k in self.format_as_float_list:
                    if k == "Rentabilitāte":
                        color = self._profitability_color_handler(v)
                        cell.fill = PatternFill(fgColor=color, fill_type= "solid")
                        cell.number_format = "0.000"
                    else:
                        cell.number_format = "0.00"
                if k in self.format_date_as_short_date:
                    cell.number_format = "DD.MM.YYYY"

                if k in self.format_as_int_list:
                    cell.number_format = "0"
                
                cell.border = self.border
                cell.value = v

    def _profitability_color_handler(self, value):
        if value < 0.75:
            color = "A30000"
        elif value < 1:
            color = "FF0000"
        elif value < 1.182:
            color = "FCD5B4"
        elif value < 1.25:
            color = "C4D79B"
        elif value < 1.40:
            color = "76933C"
        else:
            color = "538DD5"
        
        return color

    def _table_column_summary_writer(self, work_sheet= None):
        ws = work_sheet
        for column in ws.iter_cols(ws.min_column, ws.max_column):
            column_idx = column[0].col_idx
            header = work_sheet.cell(row= 1, column= column_idx).value

            summary_cell_row = int(len(column)) +2
            summary_cell_column = column[0].col_idx
            summary_cell = ws.cell(row= summary_cell_row , column=summary_cell_column)
            cell_formula_range_string = summary_cell.column_letter + str(1) + ":" + summary_cell.column_letter + str(summary_cell_row - 1)

            if header in self.format_formula_as_sum:
                summary_cell.value = "=sum(" + cell_formula_range_string + ")"
                summary_cell.number_format = "0.00"

            # atrod kolonas kuras
            if header == "Rentabilitāte":
                for column in ws.iter_cols(ws.min_column, ws.max_column):
                    column_idx = column[0].col_idx
                    column_letter = column[0].column_letter
                    header = work_sheet.cell(row= 1, column= column_idx).value
                    if header == "Nobraukti km":
                        km_column_letter  = column_letter
                    elif header == "Ienākumi":
                        revenue_column_letter = column_letter
                    elif header == "Prāmja izmaksas":
                        expenes_column_letter = column_letter
                summary_cell.value = "=" + "(" + revenue_column_letter + str(summary_cell_row) + "-" + expenes_column_letter + str(summary_cell_row) + ")" +  "/" + km_column_letter + str(summary_cell_row)
                summary_cell.number_format = "0.000"
        # for colidx in ws.iter_cols(ws.min_column, ws.max_column):
        #     print("teksts", colidx)

    

if __name__ == "__main__":
    pass