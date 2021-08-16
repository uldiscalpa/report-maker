from openpyxl import Workbook

def _as_text(value):
    if value is None:
        return ""
    return str(value)

class ExcelCreator:
    """excel report creator class"""

    def __init__(self, file_name='report.xlsx'):
        self.wb = Workbook()
        self.file_name = file_name
    
    def write_report(self, work_sheet_name= '', data_list= []):
        """Ekport data_list to init excel object"""
        ws = self.wb.create_sheet(work_sheet_name, 0)

        header_list= data_list[0].keys()
        self._header_writer(work_sheet= ws, header_list= list(header_list))
        
        self._data_writer(work_sheet= ws, data_list= data_list)
        
        self._column_adjuster(work_sheet= ws)
        self._save()
        
    def _column_adjuster(self, work_sheet):
        for column_cells in work_sheet.columns:
            length = max(len(_as_text(cell.value)) for cell in column_cells)
            if length < 10: length= 8
            work_sheet.column_dimensions[column_cells[0].column_letter].width= length
            
    def _save(self):
        self.wb.save(filename=self.file_name)
        
    def _header_writer(self, work_sheet= None, header_list= []):
        for col, header in enumerate(header_list, start= 1):
            header_cell = work_sheet.cell(column=col, row= 1)
            header_cell.value = str(header).encode(encoding='UTF-8')
    
    def _data_writer(self, work_sheet= None, data_list= []):
        for row, data in enumerate(data_list, start= 2):
            col = 0
            for k, v in data.items():
                col += 1
                cell = work_sheet.cell(column=col, row=row)
                cell.value = v
    
    def write_table_headers(self):
        pass


if __name__ == "__main__":
    pass